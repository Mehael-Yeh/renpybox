import json
from base.compat import StrEnum
from functools import partial
from typing import Any

import openpyxl
import openpyxl.styles
import openpyxl.worksheet.worksheet
from PyQt5.QtCore import QModelIndex
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QTableWidgetItem
from qfluentwidgets import TableWidget

from widget.RuleWidget import RuleWidget

class TableManager():

    class Type(StrEnum):

        GLOSSARY = "GLOSSARY"
        REPLACEMENT = "REPLACEMENT"
        TEXT_PRESERVE = "TEXT_PRESERVE"

    def __init__(self, type: str, data: list[dict[str, str]], table: TableWidget) -> None:
        super().__init__()

        self.type = type
        self.data = data
        self.table = table

        self.updating: bool = False

    def reset(self) -> None:
        self.data = []
        self.table.clearContents()
        self.table.horizontalHeader().setSortIndicator(-1, Qt.SortOrder.AscendingOrder)

    def sync(self) -> None:
        self.set_updating(True)

        dels: set[int] = set()
        for i in range(len(self.data)):
            for k in range(i + 1, len(self.data)):
                x = self.data[i]
                y = self.data[k]
                if x.get("src") == y.get("src"):
                    if x.get("dst") != "" and y.get("dst") == "":
                        dels.add(k)
                    elif x.get("dst") == "" and y.get("dst") == "" and x.get("info") != "" and y.get("info") == "":
                        dels.add(k)
                    elif x.get("dst") == "" and y.get("dst") == "" and x.get("regex") != "" and y.get("regex") == "":
                        dels.add(k)
                    else:
                        dels.add(i)
        self.data = [v for i, v in enumerate(self.data) if i not in dels]

        self.table.setRowCount(max(20, len(self.data) + 8))
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item is not None:
                    item.setText("")
                else:
                    self.table.setItem(row, col, self.generate_item(col))

        if self.type == __class__.Type.GLOSSARY:
            for row, v in enumerate(self.data):
                for col in range(self.table.columnCount()):
                    if col == 0:
                        self.table.item(row, col).setText(v.get("src", ""))
                    elif col == 1:
                        self.table.item(row, col).setText(v.get("dst", ""))
                    elif col == 2:
                        self.table.item(row, col).setText(v.get("info", ""))
                    elif col == 3:
                        rule_widget = RuleWidget(
                            show_regex = False,
                            show_case_sensitive = True,
                            case_sensitive_enabled = v.get("case_sensitive", False),
                            on_changed = partial(self._on_rule_changed, row, v),
                        )
                        self.table.setCellWidget(row, col, rule_widget)
        elif self.type == __class__.Type.REPLACEMENT:
            for row, v in enumerate(self.data):
                for col in range(self.table.columnCount()):
                    if col == 0:
                        self.table.item(row, col).setText(v.get("src", ""))
                    elif col == 1:
                        self.table.item(row, col).setText(v.get("dst", ""))
                    elif col == 2:
                        rule_widget = RuleWidget(
                            show_regex = True,
                            show_case_sensitive = True,
                            regex_enabled = v.get("regex", False),
                            case_sensitive_enabled = v.get("case_sensitive", False),
                            on_changed = partial(self._on_rule_changed, row, v),
                        )
                        self.table.setCellWidget(row, col, rule_widget)
        elif self.type == __class__.Type.TEXT_PRESERVE:
            for row, v in enumerate(self.data):
                for col in range(self.table.columnCount()):
                    if col == 0:
                        self.table.item(row, col).setText(v.get("src", ""))
                    elif col == 1:
                        self.table.item(row, col).setText(v.get("info", ""))

        self.set_updating(False)

    def export(self, path: str) -> None:
        book: openpyxl.Workbook = openpyxl.Workbook()
        sheet: openpyxl.worksheet.worksheet.Worksheet = book.active

        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 24
        sheet.column_dimensions["C"].width = 24
        sheet.column_dimensions["D"].width = 24
        sheet.column_dimensions["E"].width = 24
        TableManager.set_cell_value(sheet, 1, 1, "src", 10)
        TableManager.set_cell_value(sheet, 1, 2, "dst", 10)
        TableManager.set_cell_value(sheet, 1, 3, "info", 10)
        TableManager.set_cell_value(sheet, 1, 4, "regex", 10)
        TableManager.set_cell_value(sheet, 1, 5, "case_sensitive", 10)

        for row, item in enumerate(self.data):
            TableManager.set_cell_value(sheet, row + 2, 1, item.get("src", ""), 10)
            TableManager.set_cell_value(sheet, row + 2, 2, item.get("dst", ""), 10)
            TableManager.set_cell_value(sheet, row + 2, 3, item.get("info", ""), 10)
            TableManager.set_cell_value(sheet, row + 2, 4, item.get("regex", ""), 10)
            TableManager.set_cell_value(sheet, row + 2, 5, item.get("case_sensitive", ""), 10)

        book.save(f"{path}.xlsx")

        with open(f"{path}.json", "w", encoding = "utf-8") as writer:
            writer.write(json.dumps(self.data, indent = 4, ensure_ascii = False))

    def search(self, keyword: str, start: int) -> int:
        result: int = -1
        keyword = keyword.lower()

        for i, entry in enumerate(self.data):
            if i <= start:
                continue
            if any(keyword in v.lower() for v in entry.values() if isinstance(v, str)):
                result = i
                break

        if result == -1:
            for i, entry in enumerate(self.data):
                if i > start:
                    continue
                if any(keyword in v.lower() for v in entry.values() if isinstance(v, str)):
                    result = i
                    break

        return result

    def get_data(self) -> list[dict[str, str]]:
        return self.data

    def set_data(self, data: list[dict[str, str]]) -> None:
        self.data = data

    def get_updating(self) -> bool:
        return self.updating

    def set_updating(self, updating: bool) -> None:
        self.updating = updating

    def _on_rule_changed(self, row: int, data_ref: dict[str, str | bool], regex: bool, case_sensitive: bool) -> None:
        if self.type == __class__.Type.REPLACEMENT:
            data_ref["regex"] = regex

        data_ref["case_sensitive"] = case_sensitive

        self.table.itemChanged.emit(self.table.item(row, 0))

    def generate_item(self, col: int) -> QTableWidgetItem:
        item = QTableWidgetItem("")
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

        if self.type == __class__.Type.GLOSSARY:
            if col == 3:
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        elif self.type == __class__.Type.REPLACEMENT:
            if col == 2:
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        elif self.type == __class__.Type.TEXT_PRESERVE:
            pass

        return item

    def delete_row(self) -> None:
        selected_index = self.table.selectedIndexes()

        if selected_index == None or len(selected_index) == 0:
            return

        for row in sorted({item.row() for item in selected_index}, reverse = True):
            self.table.removeRow(row)

        self.table.itemChanged.emit(QTableWidgetItem())

    def switch_regex(self) -> None:
        selected_index: list[QModelIndex] = self.table.selectedIndexes()

        if selected_index == None or len(selected_index) == 0:
            return

        for row in {index.row() for index in selected_index}:
            item = self.table.item(row, 2)
            if item is None:
                item = QTableWidgetItem()
                self.table.setItem(row, 2, item)
            if item.text().strip() != "✅":
                item.setText("✅")
            else:
                item.setText("")

    def get_entry_by_row(self, row: int) -> dict[str, str | bool]:
        items: list[QTableWidgetItem] = [
            self.table.item(row, col)
            for col in range(self.table.columnCount())
        ]

        if self.type == __class__.Type.GLOSSARY:
            rule_widget = self.table.cellWidget(row, 3)
            case_sensitive = rule_widget.get_case_sensitive_enabled() if isinstance(rule_widget, RuleWidget) else False

            return {
                "src": items[0].text().strip() if isinstance(items[0], QTableWidgetItem) else "",
                "dst": items[1].text().strip() if isinstance(items[1], QTableWidgetItem) else "",
                "info": items[2].text().strip() if isinstance(items[2], QTableWidgetItem) else "",
                "case_sensitive": case_sensitive,
            }
        elif self.type == __class__.Type.REPLACEMENT:
            rule_widget = self.table.cellWidget(row, 2)
            regex = rule_widget.get_regex_enabled() if isinstance(rule_widget, RuleWidget) else False
            case_sensitive = rule_widget.get_case_sensitive_enabled() if isinstance(rule_widget, RuleWidget) else False

            return {
                "src": items[0].text().strip() if isinstance(items[0], QTableWidgetItem) else "",
                "dst": items[1].text().strip() if isinstance(items[1], QTableWidgetItem) else "",
                "info": "",
                "regex": regex,
                "case_sensitive": case_sensitive,
            }
        elif self.type == __class__.Type.TEXT_PRESERVE:
            return {
                "src": items[0].text().strip() if isinstance(items[0], QTableWidgetItem) else "",
                "info": items[1].text().strip() if isinstance(items[1], QTableWidgetItem) else "",
            }

    def append_data_from_table(self) -> None:
        for row in range(self.table.rowCount()):
            entry: dict[str, str | bool] = self.get_entry_by_row(row)
            if entry.get("src") != "":
                self.data.append(entry)

    def append_data_from_file(self, path: str) -> None:
        result: list[dict[str, str]] = []

        if path.lower().endswith(".json"):
            result = self.load_from_json_file(path)
        elif path.lower().endswith(".xlsx"):
            result = self.load_from_xlsx_file(path)

        self.data.extend(result)
        self.data = list({v["src"]: v for v in self.data}.values())

    def load_from_json_file(self, path: str) -> list[dict[str, str]]:
        result: list[dict[str, str]] = []

        inputs = []
        with open(path, "r", encoding = "utf-8-sig") as reader:
            inputs: dict[str, str] | list[dict[str, str]] = json.load(reader)

        if isinstance(inputs, list):
            for entry in inputs:
                if isinstance(entry, dict) == False:
                    continue
                if "src" not in entry:
                    continue

                src: str = entry.get("src", "").strip()
                if src != "":
                    result.append(
                        {
                            "src": src,
                            "dst": entry.get("dst", "").strip(),
                            "info": entry.get("info", "").strip(),
                            "regex": entry.get("regex", False),
                            "case_sensitive": entry.get("case_sensitive", False),
                        }
                    )

        if isinstance(inputs, list):
            for entry in inputs:
                if isinstance(entry, dict) == False:
                    continue
                if isinstance(entry.get("id"), int) == False:
                    continue

                id: int = entry.get("id", -1)
                name: str = entry.get("name", "").strip()
                nickname: str = entry.get("nickname", "").strip()

                if name != "":
                    result.append(
                        {
                            "src": f"\\n[{id}]",
                            "dst": name,
                            "info": "",
                            "regex": False,
                            "case_sensitive": False,
                        }
                    )
                    result.append(
                        {
                            "src": f"\\N[{id}]",
                            "dst": name,
                            "info": "",
                            "regex": False,
                            "case_sensitive": False,
                        }
                    )
                if nickname != "":
                    result.append(
                        {
                            "src": f"\\nn[{id}]",
                            "dst": name,
                            "info": "",
                            "regex": False,
                            "case_sensitive": False,
                        }
                    )
                    result.append(
                        {
                            "src": f"\\NN[{id}]",
                            "dst": name,
                            "info": "",
                            "regex": False,
                            "case_sensitive": False,
                        }
                    )

        if isinstance(inputs, dict):
            for k, v in inputs.items():
                if not isinstance(k, str):
                    continue

                src: str = k.strip()
                dst: str = str(v).strip() if v is not None else ""
                if src != "":
                    result.append(
                        {
                            "src": src,
                            "dst": dst,
                            "info": "",
                            "regex": False,
                            "case_sensitive": False,
                        }
                    )

        return result

    def load_from_xlsx_file(self, path: str) -> list[dict]:
        result: list[dict[str, str]] = []

        sheet = openpyxl.load_workbook(path).active
        for row in range(1, sheet.max_row + 1):
            data: list[str] = [
                __class__.get_cell_value(sheet, row, col)
                for col in range(1, 6)
            ]

            if len(data) == 0 or data[0] is None:
                continue

            src: str = data[0]
            dst: str = data[1]
            info: str = data[2]
            regex: bool = data[3].lower() == "true" if len(data) > 3 else False
            case_sensitive: bool = data[4].lower() == "true" if len(data) > 4 else False

            if src == "src" and dst == "dst":
                continue

            if src != "":
                result.append(
                    {
                        "src": src,
                        "dst": dst,
                        "info": info,
                        "regex": regex,
                        "case_sensitive": case_sensitive,
                    }
                )

        return result

    @classmethod
    def get_cell_value(cls, sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, column: int) -> str:
        value = sheet.cell(row = row, column = column).value

        if value is None:
            result = ""
        else:
            result = str(value)

        return result.strip()

    @classmethod
    def set_cell_value(cls, sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, column: int, value: Any, font_size: int = 9) -> None:
        if value is None:
            value = ""
        elif isinstance(value, str) and value.startswith("=") == True:
            value = "'" + value

        sheet.cell(row = row, column = column).value = value
        sheet.cell(row = row, column = column).font = openpyxl.styles.Font(size = font_size)
        sheet.cell(row = row, column = column).alignment  = openpyxl.styles.Alignment(wrap_text = True, vertical = "center", horizontal = "left")
