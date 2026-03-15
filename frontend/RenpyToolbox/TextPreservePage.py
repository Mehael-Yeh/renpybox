"""
文本保留管理页面
管理不需要翻译的文本（如专有名词、代码片段等），这些内容将在翻译过程中保持原文。
"""

from typing import List, Dict
from pathlib import Path

from PyQt5.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QFileDialog,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QAbstractItemView,
)
from PyQt5.QtCore import Qt
from qfluentwidgets import (
    CardWidget,
    PrimaryPushButton,
    PushButton,
    InfoBar,
    FluentIcon,
    TitleLabel,
    CaptionLabel,
    StrongBodyLabel,
    isDarkTheme,
    qconfig,
)

from base.Base import Base
from module.Config import Config
from base.LogManager import LogManager
from frontend.RenpyToolbox.RuleStatisticsWorker import RuleStatisticsWorker

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    load_workbook = None
    Workbook = None


class TextPreservePage(Base, QWidget):
    """文本保留管理页面"""

    HEADERS = ("原文", "备注", "命中数")
    STATS_COLUMN = 2
    STATS_COLUMN_WIDTH = 88

    def __init__(self, object_name: str, parent=None):
        Base.__init__(self)
        QWidget.__init__(self, parent)
        self.setObjectName(object_name)
        self.setProperty("toolboxPage", True)

        self.config = Config().load()
        self.logger = LogManager.get()
        self._statistics_worker = None
        self._statistics_button: PushButton | None = None
        self._statistics_snapshot_keys: List[str] = []

        self._init_ui()
        self._load_from_config()

        # 监听主题变化以更新表格配色
        qconfig.themeChanged.connect(self._on_theme_changed)

    # --- UI ---
    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(16)
        layout.setContentsMargins(24, 24, 24, 24)

        title = TitleLabel("🚫 文本保留管理")
        layout.addWidget(title)

        desc = CaptionLabel(
            "管理不需要翻译的文本（如专有名词、代码片段等），这些内容将在翻译过程中保持原文。"
        )
        desc.setWordWrap(True)
        layout.addWidget(desc)

        layout.addWidget(self._build_toolbar_card())
        layout.addWidget(self._build_table_card())
        layout.addStretch(1)

    def _build_toolbar_card(self) -> CardWidget:
        card = CardWidget(self)
        v_layout = QVBoxLayout(card)
        v_layout.setContentsMargins(16, 12, 16, 12)
        v_layout.setSpacing(6)

        # 第一排：导入/导出/保存/加载
        row1 = QHBoxLayout()
        row1.setSpacing(12)

        import_btn = PrimaryPushButton("导入 Excel", icon=FluentIcon.DOWNLOAD)
        import_btn.clicked.connect(self._on_import_excel)
        row1.addWidget(import_btn)

        export_btn = PushButton("导出 Excel", icon=FluentIcon.SHARE)
        export_btn.clicked.connect(self._on_export_excel)
        row1.addWidget(export_btn)

        save_btn = PrimaryPushButton("保存到配置", icon=FluentIcon.SAVE)
        save_btn.clicked.connect(self._save_to_config)
        row1.addWidget(save_btn)

        load_btn = PushButton("从配置加载", icon=FluentIcon.HISTORY)
        load_btn.clicked.connect(self._load_from_config)
        row1.addWidget(load_btn)

        row1.addStretch(1)
        v_layout.addLayout(row1)

        # 第二排：新增/删除/清空/重新扫描
        row2 = QHBoxLayout()
        row2.setSpacing(12)

        add_btn = PushButton("新增条目", icon=FluentIcon.ADD)
        add_btn.clicked.connect(self._add_row)
        row2.addWidget(add_btn)

        delete_btn = PushButton("删除选中", icon=FluentIcon.DELETE)
        delete_btn.clicked.connect(self._remove_selected_rows)
        row2.addWidget(delete_btn)

        dedup_btn = PushButton("去重", icon=FluentIcon.FILTER)
        dedup_btn.setToolTip("按原文去重，合并备注，优先保留有备注的行")
        dedup_btn.clicked.connect(self._deduplicate_rows)
        row2.addWidget(dedup_btn)

        clear_btn = PushButton("清空全部", icon=FluentIcon.CLOSE)
        clear_btn.setToolTip("删除所有保留文本并写入配置")
        clear_btn.clicked.connect(self._clear_all)
        row2.addWidget(clear_btn)

        statistics_btn = PushButton("统计命中", icon=FluentIcon.SEARCH)
        statistics_btn.setToolTip("基于当前 output/cache 中的缓存条目统计每条禁翻规则命中的文本数量")
        statistics_btn.clicked.connect(self._on_statistics_clicked)
        row2.addWidget(statistics_btn)
        self._statistics_button = statistics_btn

        scan_btn = PushButton("重新扫描变量", icon=FluentIcon.SYNC)
        scan_btn.setToolTip("扫描游戏目录，自动提取[variable]变量引用（清空旧数据）")
        scan_btn.clicked.connect(self._on_rescan_variables)
        row2.addWidget(scan_btn)

        row2.addStretch(1)
        v_layout.addLayout(row2)

        return card

    def _build_table_card(self) -> CardWidget:
        card = CardWidget(self)
        v_layout = QVBoxLayout(card)
        v_layout.setContentsMargins(16, 12, 16, 16)
        v_layout.setSpacing(12)

        table_label = StrongBodyLabel("保留文本列表（可直接编辑单元格）")
        v_layout.addWidget(table_label)

        self.table = QTableWidget(0, len(self.HEADERS), self)
        self.table.setHorizontalHeaderLabels(self.HEADERS)
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(self.STATS_COLUMN, QHeaderView.Fixed)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked | QAbstractItemView.EditKeyPressed)
        self.table.verticalHeader().setVisible(False)
        self.table.setColumnWidth(self.STATS_COLUMN, self.STATS_COLUMN_WIDTH)
        self.table.itemChanged.connect(self._on_table_item_changed)
        self._apply_table_theme()
        v_layout.addWidget(self.table)

        return card

    def _apply_table_theme(self) -> None:
        """根据当前主题更新表格样式"""
        if isDarkTheme():
            stylesheet = """
                QTableWidget {
                    background-color: rgb(39, 39, 39);
                    alternate-background-color: rgb(45, 45, 45);
                    color: rgb(200, 200, 200);
                    border: 1px solid rgb(55, 55, 55);
                    border-radius: 4px;
                    gridline-color: rgb(55, 55, 55);
                }
                QTableWidget::item {
                    padding: 6px;
                }
                QTableWidget::item:selected {
                    background-color: rgb(70, 70, 70);
                    color: rgb(255, 255, 255);
                }
                QHeaderView::section {
                    background-color: rgb(50, 50, 50);
                    color: rgb(200, 200, 200);
                    padding: 8px;
                    border: none;
                    border-bottom: 1px solid rgb(65, 65, 65);
                    font-weight: bold;
                }
            """
        else:
            stylesheet = """
                QTableWidget {
                    background-color: rgb(255, 255, 255);
                    alternate-background-color: rgb(248, 248, 248);
                    color: rgb(32, 32, 32);
                    border: 1px solid rgb(220, 220, 220);
                    border-radius: 4px;
                    gridline-color: rgb(230, 230, 230);
                }
                QTableWidget::item {
                    padding: 6px;
                }
                QTableWidget::item:selected {
                    background-color: rgb(210, 210, 210);
                    color: rgb(0, 0, 0);
                }
                QHeaderView::section {
                    background-color: rgb(245, 245, 245);
                    color: rgb(32, 32, 32);
                    padding: 8px;
                    border: none;
                    border-bottom: 1px solid rgb(220, 220, 220);
                    font-weight: bold;
                }
            """
        self.table.setStyleSheet(stylesheet)

    def _on_theme_changed(self) -> None:
        """主题切换时同步更新表格样式"""
        self._apply_table_theme()

    def _create_table_item(self, text: str = "", *, editable: bool = True) -> QTableWidgetItem:
        item = QTableWidgetItem(text)
        if not editable:
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            item.setTextAlignment(Qt.AlignCenter)
        return item

    def _build_statistics_entry_key(self, item: Dict[str, str]) -> str:
        return self._normalize_src(str(item.get("src", "") or ""))

    def _set_statistics_buttons_enabled(self, enabled: bool) -> None:
        if self._statistics_button is not None:
            self._statistics_button.setEnabled(enabled)

    def _invalidate_statistics(self) -> None:
        self._statistics_snapshot_keys = []
        self.table.blockSignals(True)
        try:
            for row in range(self.table.rowCount()):
                item = self.table.item(row, self.STATS_COLUMN)
                if item is None:
                    item = self._create_table_item("", editable=False)
                    self.table.setItem(row, self.STATS_COLUMN, item)
                else:
                    item.setText("")
        finally:
            self.table.blockSignals(False)

    def _on_table_item_changed(self, item: QTableWidgetItem) -> None:
        if item is None:
            return
        if item.column() != 0:
            return
        self._invalidate_statistics()

    # --- 数据操作 ---
    def _add_row(self):
        row = self.table.rowCount()
        self.table.insertRow(row)
        self.table.setItem(row, 0, self._create_table_item(""))
        self.table.setItem(row, 1, self._create_table_item(""))
        self.table.setItem(row, self.STATS_COLUMN, self._create_table_item("", editable=False))
        self.table.setCurrentCell(row, 0)
        self._invalidate_statistics()

    def _remove_selected_rows(self):
        row = self.table.currentRow()
        if row < 0:
            InfoBar.warning("提示", "请选择需要删除的条目", parent=self)
            return
        self.table.removeRow(row)
        self._invalidate_statistics()

    def _deduplicate_rows(self):
        """按原文去重，优先保留有备注的条目"""
        entries = self._collect_table_data()
        if not entries:
            InfoBar.info("提示", "表格为空，暂无可去重的数据", parent=self)
            return

        key_index: Dict[str, int] = {}
        deduped: List[Dict[str, str]] = []
        for item in entries:
            key = self._normalize_src(item.get("src", ""))
            if not key:
                continue
            if key not in key_index:
                deduped.append({"src": item.get("src", "").strip(), "comment": item.get("comment", "").strip()})
                key_index[key] = len(deduped) - 1
            else:
                existing = deduped[key_index[key]]
                merged = self._merge_entries(existing, item)
                deduped[key_index[key]] = merged

        removed = len(entries) - len(deduped)
        if removed > 0:
            self._set_table_data(deduped)
            InfoBar.success("完成", f"已去除重复 {removed} 条，保留 {len(deduped)} 条", parent=self)
        else:
            InfoBar.info("提示", "未发现重复条目", parent=self)

    def _clear_all(self):
        """清空表格并写回配置"""
        self.table.setRowCount(0)
        self.config = Config().load()
        self.config.text_preserve_data = []
        self.config.text_preserve_enable = False
        self.config.save()
        self._invalidate_statistics()
        InfoBar.success("已清空", "已删除所有保留文本并写入配置", parent=self)

    def _load_from_config(self):
        data = getattr(self.config, "text_preserve_data", []) or []
        converted = []
        for item in data:
            if isinstance(item, dict):
                converted.append(
                    {
                        "src": item.get("src", ""),
                        "comment": item.get("comment", item.get("info", "")),
                    }
                )
            elif isinstance(item, str): # 兼容旧格式或纯字符串列表
                converted.append(
                    {
                        "src": item,
                        "comment": "",
                    }
                )
        self._set_table_data(converted)
        InfoBar.success("完成", f"已从配置加载 {len(converted)} 条保留文本", parent=self)

    def _save_to_config(self):
        entries = self._collect_table_data()
        self.config = Config().load()
        self.config.text_preserve_data = entries
        self.config.text_preserve_enable = True if entries else self.config.text_preserve_enable
        self.config.save()
        InfoBar.success("保存成功", f"已写入 {len(entries)} 条保留文本到配置", parent=self)

    def _on_import_excel(self):
        if load_workbook is None:
            InfoBar.error("错误", "未安装 openpyxl，无法导入 Excel", parent=self)
            return

        path, _ = QFileDialog.getOpenFileName(
            self,
            "选择 Excel 文件",
            "",
            "Excel 文件 (*.xlsx)"
        )
        if not path:
            return
        try:
            workbook = load_workbook(path)
            sheet = workbook.active
            headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
            header_map = self._build_header_map(headers)
            if "src" not in header_map:
                raise ValueError("未找到“原文”列，请确认模板。")

            items: List[Dict[str, str]] = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                src = self._safe_cell(row, header_map.get("src"))
                comment = self._safe_cell(row, header_map.get("comment"))
                if not src:
                    continue
                items.append({"src": src, "comment": comment})

            self._set_table_data(items)
            InfoBar.success("导入成功", f"已导入 {len(items)} 条保留文本", parent=self)
        except Exception as e:
            self.logger.error(f"导入失败: {e}")
            InfoBar.error("错误", f"导入失败: {e}", parent=self)

    def _on_export_excel(self):
        if Workbook is None:
            InfoBar.error("错误", "未安装 openpyxl，无法导出 Excel", parent=self)
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            "保存 Excel",
            "",
            "Excel 文件 (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        entries = self._collect_table_data()
        if not entries:
            InfoBar.warning("提示", "当前表格为空，未导出文件", parent=self)
            return

        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "TextPreserve"
            sheet.append(list(self.HEADERS[:-1]))
            for item in entries:
                sheet.append([item.get("src", ""), item.get("comment", "")])
            workbook.save(path)
            InfoBar.success("导出成功", f"已保存到 {path}", parent=self)
        except Exception as e:
            self.logger.error(f"导出失败: {e}")
            InfoBar.error("错误", f"导出失败: {e}", parent=self)

    def _on_statistics_clicked(self) -> None:
        if self._statistics_worker and self._statistics_worker.isRunning():
            InfoBar.info("提示", "命中统计正在进行中，请稍候…", parent=self)
            return

        entries = self._collect_table_data()
        if not entries:
            InfoBar.info("提示", "当前禁翻表为空，暂无可统计的数据", parent=self)
            return

        config = Config().load()
        self._statistics_snapshot_keys = [
            self._build_statistics_entry_key(entry) for entry in entries
        ]
        self._set_statistics_buttons_enabled(False)

        worker = RuleStatisticsWorker(
            mode = RuleStatisticsWorker.MODE_TEXT_PRESERVE,
            config = config,
            entries = entries,
            parent = self,
        )
        worker.finished.connect(self._on_statistics_finished)
        self._statistics_worker = worker

        worker.start()

    def _on_statistics_finished(self, success: bool, message: str, payload) -> None:
        self._set_statistics_buttons_enabled(True)

        worker = self._statistics_worker
        self._statistics_worker = None
        if worker is not None:
            worker.deleteLater()

        if success == False:
            InfoBar.warning("统计失败", message, parent=self)
            return

        if not isinstance(payload, dict):
            InfoBar.warning("统计失败", "统计结果格式无效", parent=self)
            return

        counts = payload.get("counts", [])
        if not isinstance(counts, list):
            InfoBar.warning("统计失败", "统计结果缺少命中数", parent=self)
            return

        current_entries = self._collect_table_data()
        current_keys = [self._build_statistics_entry_key(entry) for entry in current_entries]
        if current_keys != self._statistics_snapshot_keys:
            self._invalidate_statistics()
            InfoBar.warning("提示", "禁翻表内容已变化，请重新执行一次统计", parent=self)
            return

        self.table.blockSignals(True)
        try:
            for row, count in enumerate(counts):
                if row >= self.table.rowCount():
                    break

                item = self.table.item(row, self.STATS_COLUMN)
                if item is None:
                    item = self._create_table_item("", editable=False)
                    self.table.setItem(row, self.STATS_COLUMN, item)
                item.setText(str(max(0, int(count))))
        finally:
            self.table.blockSignals(False)

        counted_item_total = int(payload.get("counted_item_total", 0))
        InfoBar.success(
            "统计完成",
            f"已统计 {len(counts)} 条禁翻规则，样本条目 {counted_item_total} 条",
            parent=self,
        )

    # --- 工具方法 ---
    def _set_table_data(self, items: List[Dict[str, str]]):
        self.table.blockSignals(True)
        try:
            self.table.setRowCount(0)
            for item in items:
                row = self.table.rowCount()
                self.table.insertRow(row)
                self.table.setItem(row, 0, self._create_table_item(item.get("src", "")))
                self.table.setItem(row, 1, self._create_table_item(item.get("comment", "")))
                self.table.setItem(row, self.STATS_COLUMN, self._create_table_item("", editable=False))
        finally:
            self.table.blockSignals(False)
        self._invalidate_statistics()

    def _collect_table_data(self) -> List[Dict[str, str]]:
        results: List[Dict[str, str]] = []
        rows = self.table.rowCount()
        for row in range(rows):
            src_item = self.table.item(row, 0)
            comment_item = self.table.item(row, 1)
            src = (src_item.text() if src_item else "").strip()
            comment = (comment_item.text() if comment_item else "").strip()
            if not src:
                continue
            results.append({"src": src, "comment": comment, "info": comment})
        return results

    @staticmethod
    def _normalize_src(text: str) -> str:
        if not text:
            return ""
        return text.strip().strip("\"'“”‘’").lower()

    @staticmethod
    def _merge_entries(base: Dict[str, str], incoming: Dict[str, str]) -> Dict[str, str]:
        def _clean(v: str) -> str:
            return v.strip() if isinstance(v, str) else ""

        merged = {"src": _clean(base.get("src")), "comment": _clean(base.get("comment"))}
        incoming_clean = {"src": _clean(incoming.get("src")), "comment": _clean(incoming.get("comment"))}

        # 保留有备注的
        if incoming_clean["comment"]:
            if not merged["comment"] or len(incoming_clean["comment"]) > len(merged["comment"]):
                merged["comment"] = incoming_clean["comment"]

        if incoming_clean["src"] and not merged["src"]:
            merged["src"] = incoming_clean["src"]
        return merged

    @staticmethod
    def _build_header_map(headers: List[str]) -> Dict[str, int]:
        alias = {
            "src": {"原文", "原始文本", "source", "src", "text"},
            "comment": {"备注", "说明", "comment", "note", "备注信息"},
        }
        mapping = {}
        for index, name in enumerate(headers):
            lower_name = name.lower()
            for key, options in alias.items():
                if lower_name in {opt.lower() for opt in options} and key not in mapping:
                    mapping[key] = index
        return mapping

    @staticmethod
    def _safe_cell(row, index: int) -> str:
        if index is None:
            return ""
        if index >= len(row):
            return ""
        value = row[index]
        return "" if value is None else str(value).strip()

    @staticmethod
    def _list_scan_candidates(config: Config) -> List[Path]:
        """根据当前配置推断变量扫描候选目录（按优先级排序并去重）"""
        raws = [
            getattr(config, "input_folder", ""),
            getattr(config, "output_folder", ""),
            getattr(config, "renpy_game_folder", ""),
        ]
        candidates: List[Path] = []
        for raw in raws:
            if not raw:
                continue
            path = Path(raw)
            if not path.exists():
                continue

            if path.is_file():
                if path.suffix.lower() == ".rpy":
                    candidates.append(path.parent)
                continue

            # 若选择的是 tl 目录，优先回退到上层 game 目录
            if path.name.lower() == "tl" and path.parent.exists():
                candidates.append(path.parent)

            game_child = path / "game"
            if game_child.exists() and game_child.is_dir():
                candidates.append(game_child)

            candidates.append(path)

        deduped: List[Path] = []
        seen: set[str] = set()
        for p in candidates:
            try:
                key = str(p.resolve()).lower()
            except Exception:
                key = str(p).lower()
            if key in seen:
                continue
            seen.add(key)
            deduped.append(p)
        return deduped

    @staticmethod
    def _count_rpy_files_without_tl(root: Path) -> int:
        """统计目录下可用于扫描变量的 rpy 数量（排除 tl 目录）"""
        count = 0
        for rpy_file in root.rglob("*.rpy"):
            if "tl" in [part.lower() for part in rpy_file.parts]:
                continue
            count += 1
        return count

    def _on_rescan_variables(self):
        """重新扫描游戏目录，提取[variable]变量引用到禁翻表（清空旧数据）"""
        import re
        
        # 重新加载配置以获取最新目录配置
        self.config = Config().load()

        candidates = self._list_scan_candidates(self.config)
        if not candidates:
            InfoBar.warning("警告", "未找到可扫描目录，请先设置输入/输出目录或游戏目录", parent=self)
            return

        # 按候选优先级选择第一个可扫描目录：
        # input/output 优先于 renpy_game_folder，避免跳到历史项目。
        game_path = None
        fallback_path = None
        for candidate in candidates:
            if fallback_path is None:
                fallback_path = candidate
            try:
                count = self._count_rpy_files_without_tl(candidate)
            except Exception:
                count = -1
            if count > 0:
                game_path = candidate
                break

        if game_path is None:
            game_path = fallback_path

        if game_path is None or not game_path.exists():
            InfoBar.error("错误", "无法确定扫描目录", parent=self)
            return

        # 正则匹配 [variable_name]
        RE_VARIABLE_IN_TEXT = re.compile(r'\[([\w.]+)\]')
        
        found_preserves = set()
        try:
            for rpy_file in game_path.rglob("*.rpy"):
                # 跳过 tl 目录，避免将翻译产物中的占位污染禁翻表
                if "tl" in [part.lower() for part in rpy_file.parts]:
                    continue
                try:
                    content = rpy_file.read_text(encoding="utf-8", errors="ignore")
                    var_matches = RE_VARIABLE_IN_TEXT.findall(content)
                    for var_name in var_matches:
                        found_preserves.add(f"[{var_name}]")
                except Exception:
                    pass
        except Exception as e:
            InfoBar.error("错误", f"扫描失败: {e}", parent=self)
            return
        
        if not found_preserves:
            # 清空禁翻表
            self.config.text_preserve_data = []
            self.config.text_preserve_enable = False
            self.config.save()
            self._load_from_config()
            InfoBar.info("提示", f"未找到变量引用，已清空禁翻表（扫描目录：{game_path}）", parent=self)
            return
        
        # 完全清空旧数据，只保留新扫描的 [variable]
        new_preserves = []
        for text in sorted(found_preserves):
            new_preserves.append({"src": text})
        
        # 保存到配置
        self.config.text_preserve_data = new_preserves
        self.config.text_preserve_enable = True
        self.config.save()
        
        # 刷新表格
        self._load_from_config()
        
        InfoBar.success("完成", f"已扫描到 {len(new_preserves)} 个变量引用（扫描目录：{game_path}）", parent=self)
